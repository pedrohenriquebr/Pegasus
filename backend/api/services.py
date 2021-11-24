from typing import Any, List

from datetime import datetime
from api.uof import Uof
from sheetsorm.orm import SheetsORM
from database.entities import TBL_Transactions,TBL_Account,TBL_AccountGroup,TBL_Category
from helpers.date import DEFAULT_DATE_FORMAT
from database.schema import schema
from api.config import *
import adapters.bank_statement.inter

from openpyxl.worksheet.table import Table
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.colors import Color

import pandas as pd
import numpy as np

from api.models import ImportationCommand

class TransactionsService:

    def __init__(self, uof: Uof):
        self.uof = uof
        self.repo = uof.TransactionsRepository


    def get_all_transactions(self,id_account, page=0, limit=10):
        accounts = self.uof.AccountRepository.get_all()
        categories = self.uof.CategoryRepository.get_all()
        data = [self._format_transaction(row,accounts,categories) for row in self.repo.find(lambda x: x.ID_Account == id_account)]
        return {
            'count': len(data),
            'data': data[page*limit:page*limit+limit] if limit != -1 else data
        }

    def _format_transaction(self, transaction: TBL_Transactions, accounts: List[TBL_Account], categories: List[TBL_Category]):
        category = [c for c in categories if c.ID_Category == transaction.ID_Category]
        account = [a for a in accounts if a.ID_Account == transaction.ID_Account]
        account_destination = [a for a in accounts if a.ID_Account == transaction.ID_AccountDestination]
        return {
            'id': transaction.ID_Transaction,
            'transaction_date': transaction.DT_TransactionDate.strftime(DEFAULT_DATE_FORMAT) if transaction.DT_TransactionDate else '',
            'registration_date': transaction.DT_RegistrationDate.strftime(DEFAULT_DATE_FORMAT) if transaction.DT_RegistrationDate else '',
            'type': transaction.CD_Type,
            'amount': transaction.NR_Amount,
            'description': transaction.DS_Description,
            'category':  category[0].DS_Name if len(category) > 0 else 'Desconhecido',
            'account_destination': account_destination[0].DS_Name if len(account_destination) > 0 else '',
            'account': account[0].DS_Name if len(account) > 0 else '',
            'id_category': transaction.ID_Category,
            'id_account': transaction.ID_Account,
            'id_account_destination': transaction.ID_AccountDestination,
            'is_imported': transaction.IC_Imported,
            'imported_date': transaction.DT_ImportedDate.strftime(DEFAULT_DATE_FORMAT) if transaction.DT_ImportedDate else '',
            'balance': transaction.NR_Balance
        }


class AccountsService:
    def __init__(self, uof: Uof):
        self.uof = uof
        self.repo = self.uof.AccountRepository
    
    def get_all_accounts(self,offset=0, limit=100):
        accounts_groups = self.uof.AccountGroupRepository.get_all()
        data = [self._format_account(row, accounts_groups) for row in self.repo.get_all()]
        return {
                'count': len(data),
                'data': data[offset*limit:offset*limit+limit]
            }
    
    def get_all_accounts_names(self):
        data = [{'id': row.ID_Account, 'name': row.DS_Name} for row in self.repo.get_all()]
        return {
            'count': len(data),
            'data': data
        }
    
    def _format_account(self, account: TBL_Account, accounts_groups: list):
        group = [group.ID_AccountGroup == account.ID_AccountGroup for group in accounts_groups]
        return {
            'id': account.ID_Account,
            'name': account.DS_Name,
            'group': ( group[0].DS_Name if group[0].DS_Name else '') if len(group) > 0 else '',
            'created_date': account.DT_CreatedDate.strftime(DEFAULT_DATE_FORMAT) if account.DT_CreatedDate else '',
            'initial_amount': account.NR_InitialAmount,
            'number': account.NR_Number,
            'bankoffice_number': account.NR_BankOfficeNumber,
            'bank': account.DS_Bank
        }


class CategoriesService:
    def __init__(self, uof):
        self.uof = uof 
        self.repo = self.uof.CategoryRepository
    
    def get_all_categories(self,offset=0, limit=100):
        data = [self._format_category(row) for row in self.repo.get_all()]
        return {
                'count': len(data),
                'data': data[offset*limit:offset*limit+limit]
            }
    
    def get_all_categories_names(self):
        categories  = self.repo.get_all()
        data = [self._format_category(row,categories) for row in categories]
        return {
            'count': len(data),
            'data': data
        }
    
    def _format_category(self, category: TBL_Category, categories: List[TBL_Category]):
        cat = [cat for cat in categories if cat.ID_Category == category.ID_CategoryParent ]
        return {
            'id': category.ID_Category,
            'name': category.DS_Name,
            'parent': ( cat[0].DS_Name if cat[0].DS_Name else '') if len(cat) > 0 else '',
            'id_parent': category.ID_CategoryParent if category.ID_CategoryParent else 0,
            'level': category.NR_Level,
        }


class ExportationService:
    def  __init__(self, uof: Uof):
        self.uof = uof
        self.sheets = {}
        self.excel_name_file = 'output.xlsx' 
        self.excel_file = os.path.abspath(UPLOAD_FOLDER + '/' + self.excel_name_file)
        
    def export_data(self, id_account:int=0, year:int=0, month=0):
        self.TBL_Category = self.uof.CategoryRepository.to_dataframe()
        self.TBL_DescriptionCategory = self.uof.DescriptionCategoryRepository.to_dataframe()
        self.DIM_NR_Income = schema['DIM_NR_Income']
        self.TBL_Transactions = self.uof.TransactionsRepository.to_dataframe()
        self.accounts =  self.uof.AccountRepository.get_all() \
                                            if id_account == 0 \
                                            else self.uof.AccountRepository.find(lambda x: x.ID_Account == id_account)
        if year == 0:
            year = datetime.now().year
        
        self.base = self.load_data(self.TBL_Transactions)
        self.categorias  = self._load_categories()
        self.historico_df = self._load_historic()

        self.base  = self.base[self.base['DT_TransactionDate'].dt.year == year]

        if month >  0:
            self.base  = self.base[self.base['DT_TransactionDate'].dt.month == month]
        

        
        for account in self.accounts:
            print(f'Generating sheet for account "{account.DS_Name}"')
            self.sheets[account.DS_Name] = self._transform(self.base,account.ID_Account,account.NR_InitialAmount)

        # self.export_csv()
        self.export_xlsx()
        return self.excel_name_file

    def apply_extra_styles(self,excel_file: str, report) -> None:
        # load workbook template.xlsx from current directory
        wb = load_workbook(excel_file)

        ws = wb[report]

        # esconde a primeira coluna
        ws.column_dimensions['A'].hidden = True

        # esconder a primeira linha
        ws.row_dimensions[1].hidden = True

        # esconder a 4ª linha
        ws.row_dimensions[4].hidden = True

        # ajustar fonte dos nomes dos meses
        for row in ws['B2:BV2']:
            for cell in row:
                cell.font = Font(size=20)

        # ajustar fonte e cor de fundo dos numeros das semanas
        for row in ws['C3:BV3']:
            for cell in row:
                # converte string em 
                cell.fill = PatternFill("solid", fgColor=Color(rgb='00305496'))
                cell.font = Font(color="ffffffff", bold=True, size=14)

        # percorre as colunas da sheet
        for col in ws.columns:
            # auto ajusta a largura da coluna
            col_name = get_column_letter(col[0].column)
            if col_name == 'B':
                ws.column_dimensions[col_name].width = 38.57
            else:
                ws.column_dimensions[col_name].auto_size = True

        # alterar zoom da worksheet para 85%
        ws.sheet_view.zoomScale = 85

        # formatar valores para o formato R$ ##,## da célula C5:BJ45
        for row in ws['C5:BV43']:
            for cell in row:
                cell.number_format = 'R$ #,##0.00'
                cell.font = Font(size=14)

        # agrupar linhas
        last_start  = 9
        last_end = 8
        for cell in ws['B9:B47']:
            # verificar se o valor da célula é NoneType
            if cell[0].value is None:
                continue
            if cell[0].value.strip() == 'TOTAL':
                continue
            
            if cell[0].value.strip() in [*self.cat_pais, 'Desconhecido']:
                ws.row_dimensions.group(last_start,last_end, hidden=True)
                last_start  = cell[0].row + 2
                last_end = last_start - 1
            else:
                last_end +=1

        ws.row_dimensions.group(last_start,last_end, hidden=True)

        ws.freeze_panes = ws['C2']

        wb[report].sheet_properties.tabColor = '203764'


        wb.save(self.excel_file)
        wb.close()
      
    def export_xlsx(self, ) -> None:
        with pd.ExcelWriter(self.excel_file, mode='w') as writer:
            for name, df in self.sheets.items():
                self.get_report_styles(df).to_excel(writer,sheet_name=name)
            
            print(f'Exporting to {self.excel_file} sheets')
        for name, df in self.sheets.items():
            self.apply_extra_styles(self.excel_file, name)
   
    def get_report_styles(self,pivot_table):
        # the  Styler object dont apply styles on columns or indexes
        
        styles = {
            'TOTAL': 'background-color: #8EA9DB; color: black; font-weight: bold',
            'NR_Balance': 'font-weight: bold;background-color: white;color:black',
            'RECEITA': 'font-weight: bold;background-color: white;color:black',
            'Desconhecido': 'background-color: #305496; color: white; font-weight: bold'
        }

        return pivot_table.style.apply(lambda value, props='': np.where(pivot_table['CATEGORIA'] == 'TOTAL', props, ''), props=styles['TOTAL'], axis=0)\
            .apply(lambda value, props='': np.where(pivot_table['CATEGORIA'] == 'NR_Balance', props, ''), props=styles['NR_Balance'], axis=0)\
            .apply(lambda value, props='': np.where(pivot_table['CATEGORIA'] == 'RECEITA', props, ''), props=styles['RECEITA'], axis=0)\
            .apply(lambda value, props='': np.where(pivot_table['CATEGORIA'] == 'Desconhecido', props, ''), props=styles['Desconhecido'], axis=0)\
            .apply(lambda value, props='': np.where(pivot_table['CATEGORIA'].isin(self.cat_pais), props, ''), props='background-color: #305496; color:white; font-weight:bold', axis=0)\
            .apply(lambda value, props='': np.where(pivot_table['CATEGORIA'].isin(self.cats), props, ''), props='background-color: white; color:black; font-weight:bold', axis=0)\
            .set_properties(**{'border-color':'Black','border-width':'thin','border-style':'solid'})\
            .apply(self._apply_styles, props='font-size: 14pt', col=('CATEGORIA', '', ''),axis=0)

    def _apply_styles(self, x, col, props):
        return [ props if x.name == col else 'font-size: 11pt' for v in x.index ]
    
    def export_csv(self) -> None:
        if not os.path.exists(CACHE_DIR):    
            os.mkdir(CACHE_DIR)
        print('Export DW')

        for outfile, df in self.sheets.items():
            print(f'Exporting {outfile}.csv')
            df.to_csv(os.path.join(CACHE_DIR, f'{outfile}.csv'),sep=';',index=False, header=True, )
    
    def _find_keys(self,dict_):
        key_list = []
        def recursive(obj, last_parent=None,level=1):
            for key in obj.keys():
                if type(obj[key]) == dict:
                    key_list.append({'cat':key, 'level':level , 'cat_parent': last_parent, 'value': None})
                    recursive(obj[key], last_parent=key,level=level+1)
                else:
                    a = len(obj[key])
                    if a >0:
                        for value in obj[key]:
                            key_list.append({'cat': key,'cat_parent': last_parent, 'level': level, 'value': value })

                    else:
                        key_list.append({'cat': key,'cat_parent': last_parent, 'level': level, 'value': None})
            
        recursive(dict_)
        return sorted(key_list, key=lambda x: x['level'])

    def load_data(self,transactions: pd.DataFrame) -> pd.DataFrame:
        base = pd.DataFrame({
            'ID_Base': pd.Series(dtype='int',data=list(range(1,transactions.shape[0]))),
            'ID_Account': transactions['ID_Account'],
            'ID_Category': transactions['ID_Category'],
            'ID_AccountDestination': transactions['ID_AccountDestination'],
            'CD_Type': transactions['CD_Type'],
            'DT_TransactionDate': transactions['DT_TransactionDate'],
            'DS_Description': transactions['DS_Description'],
            'NR_Value': transactions['NR_Amount'],
            'NR_Balance': transactions['NR_Balance'],
        })
        
        base['DT_TransactionDate'] = pd.to_datetime(base['DT_TransactionDate'])
        return base.sort_values('DT_TransactionDate')

        
    def _transform(self, base, id_account, initial_amount):
        # Dimensions
        self.base_df = base[(self.base['ID_Account'] == id_account) | (base['ID_AccountDestination'] == id_account )]

        # dim_entrada
        self.DIM_NR_Income = self.base_df[(self.base_df['CD_Type'] == 'Income') \
                                            | ( (self.base_df['CD_Type'] == 'Transfer') \
                                                & (self.base_df['ID_AccountDestination'] == id_account) )]['NR_Value']
        self.DIM_NR_Income.drop_duplicates(inplace=True)
        self.DIM_NR_Income.reset_index(drop=True,inplace=True)
        self.DIM_NR_Income = self.DIM_NR_Income.to_frame()
        self.DIM_NR_Income['ID_Income'] = self.DIM_NR_Income.index + 1
        self.DIM_NR_Income['ID_Income'] = self.DIM_NR_Income['ID_Income'].astype(np.int64)

        # dim_saida
        self.DIM_NR_Expense = self.base_df[(self.base_df['NR_Value'] == 'Expense') \
                                             | ((self.base_df['CD_Type'] == 'Transfer') \
                                                & (self.base_df['ID_Account'] == id_account))]['NR_Value']
        self.DIM_NR_Expense.drop_duplicates(inplace=True)
        self.DIM_NR_Expense.reset_index(drop=True,inplace=True)
        self.DIM_NR_Expense = self.DIM_NR_Expense.to_frame()
        self.DIM_NR_Expense['ID_Expense'] = self.DIM_NR_Expense.index + 1
        self.DIM_NR_Expense['ID_Expense'] = self.DIM_NR_Expense['ID_Expense'].astype(np.int64)

        # dim_historico
        self.DIM_DS_Description = self.base_df['DS_Description']
        self.DIM_DS_Description.drop_duplicates(inplace=True)
        self.DIM_DS_Description.reset_index(drop=True,inplace=True)
        self.DIM_DS_Description = self.DIM_DS_Description.to_frame()
        self.DIM_DS_Description['ID_DS_Description'] = self.DIM_DS_Description.index + 1
        self.DIM_DS_Description['ID_DS_Description'] = self.DIM_DS_Description['ID_DS_Description'].astype(np.int64)

        # dim_lancamento
        self.DIM_DT_TransactionDate = self.base_df['DT_TransactionDate']
        self.DIM_DT_TransactionDate.drop_duplicates(inplace=True)
        self.DIM_DT_TransactionDate.reset_index(drop=True,inplace=True)
        self.DIM_DT_TransactionDate = self.DIM_DT_TransactionDate.to_frame()
        self.DIM_DT_TransactionDate['ID_TransactionDate'] = self.DIM_DT_TransactionDate.index + 1
        self.DIM_DT_TransactionDate['ID_TransactionDate'] = self.DIM_DT_TransactionDate['ID_TransactionDate'].astype(np.int64)

        # dim_saldo
        self.DIM_NR_Balance = self.base_df['NR_Balance']
        self.DIM_NR_Balance.drop_duplicates(inplace=True)
        self.DIM_NR_Balance.reset_index(drop=True,inplace=True)
        self.DIM_NR_Balance = self.DIM_NR_Balance.to_frame()
        self.DIM_NR_Balance['ID_Balance'] = self.DIM_NR_Balance.index + 1
        self.DIM_NR_Balance['ID_Balance'] = self.DIM_NR_Balance['ID_Balance'].astype(np.int64)


        self.merged = pd.merge(self.base_df,self.DIM_DT_TransactionDate, how='left',on='DT_TransactionDate')
        self.merged = pd.merge(self.merged,self.DIM_NR_Income, how='left',on='NR_Value')
        self.merged = pd.merge(self.merged,self.DIM_NR_Expense, how='left',on='NR_Value')
        self.merged = pd.merge(self.merged,self.DIM_DS_Description, how='left',on='DS_Description')
        self.merged = pd.merge(self.merged,self.DIM_NR_Balance, how='left',on='NR_Balance')

        # # Remove columns
        self.merged.drop(columns=['DS_Description', 'NR_Value', 'DT_TransactionDate','NR_Balance'], inplace=True,errors='ignore')
        self.DIM_NR_Expense['NR_Value'] = self.DIM_NR_Expense['NR_Value'].abs()

        self.cat_pais = self.TBL_Category[self.TBL_Category['NR_Level'] == 1]['DS_Name'].unique()
        self.cats = self.TBL_Category[self.TBL_Category['NR_Level'] > 1]['DS_Name'].unique()

        report = self._build_pivot_table(self.merged, initial_amount)
        return report
        
        
       
    def _build_pivot_table(self, fat_table: pd.DataFrame, initial_amount: float) -> pd.DataFrame:
        report  = self.build_weekly_report(fat_table)
        report = report[['DT_TransactionDate','ANO','MÊS','SEMANA','ENTRADA','SAIDA','CATEGORIA']]\
            .groupby(['DT_TransactionDate','ANO','MÊS','SEMANA','CATEGORIA'])\
            .sum([['ENTRADA','SAIDA']]).reset_index().drop(columns=['DT_TransactionDate'])
        filtro = report
        filtro = pd.merge(filtro, self.TBL_Category, how='left',  left_on='CATEGORIA', right_on='DS_Name')
        filtro = pd.merge(filtro, self.TBL_Category, how='left',  left_on='ID_CategoryParent', right_on='ID_Category')
        filtro['CATEGORIA_PAI'] = filtro['DS_Name_y']
        filtro =  filtro[['ANO', 'MÊS','SEMANA', 'CATEGORIA', 'CATEGORIA_PAI','ENTRADA', 'SAIDA']]
        filtro['NR_Value'] = filtro['ENTRADA'] + filtro['SAIDA']
        filtro['CATEGORIA_PAI']  = filtro['CATEGORIA_PAI'].fillna('Desconhecido')
        pivot_table = pd.pivot_table(filtro, values=['NR_Value'],index=['CATEGORIA_PAI','CATEGORIA'],columns=['MÊS','SEMANA'], aggfunc=(np.sum))

        d = filtro.groupby(['CATEGORIA_PAI','CATEGORIA'],as_index=True).sum().index

        groups =[]
        k = []
        for row in d:
            if row[0] not in groups:
                groups.append(row[0])
                k.append((row[0],row[0]))
                k.append((row[0],'TOTAL'))
            k.append((row[0],row[1]))

        pivot_table = pivot_table.reindex(pd.MultiIndex.from_tuples(k,names=['CATEGORIA_PAI','CATEGORIA']))
        pivot_table = pivot_table.reset_index()

        pivot_table.loc[-1] = ['' , 'RECEITA', *[self.calc_income(report,pivot_table,idx) for idx in range(2,pivot_table.shape[1])]]  
        pivot_table.index = pivot_table.index + 1 
        balance_values = []

        for idx in range(2,pivot_table.shape[1]):
            initial_amount = self.calc_balance(initial_amount, report,pivot_table,idx)
            balance_values.append(initial_amount)

        pivot_table.loc[-1] = ['' , 'SALDO', *balance_values]  
        pivot_table.index = pivot_table.index + 1  
        pivot_table = pivot_table.sort_index()
        total_search = pivot_table[pivot_table['CATEGORIA'] == 'TOTAL'] 
        rows =  total_search.iterrows()

        series  = []
        for row in rows:
            cat  = row[1][0]
            label  = row[1][1]
            values = []
            for idx in range(2,pivot_table.shape[1]):
                month, week = pivot_table.columns[idx][1], pivot_table.columns[idx][2]
                values = [*values, filtro[(filtro['MÊS'] == month ) & (filtro['SEMANA'] == week ) & (filtro['CATEGORIA_PAI'] == cat)]['NR_Value'].sum()]
            series.append([cat, label, *values])

        pivot_table[pivot_table['CATEGORIA'] == 'TOTAL']  = pd.DataFrame(series, columns=total_search.columns, index=total_search.index)
        pivot_table = pivot_table.drop(columns=['CATEGORIA_PAI'])
        return pivot_table 
    
    
    def calc_income(self, report, pivot_table: pd.DataFrame, idx: int):
        # MÊS , SEMANA 
        month, week = pivot_table.columns[idx][1], pivot_table.columns[idx][2]
        temp2  = report[(report['MÊS'] == month) & (report['SEMANA'] == week)]
        return round(temp2['ENTRADA'].sum(),10)

    def calc_balance(self, init_balance, origin_report, pivot_table: pd.DataFrame, idx: int):
        # MÊS , SEMANA 
        month, week = pivot_table.columns[idx][1], pivot_table.columns[idx][2]

        temp2  = origin_report[(origin_report['MÊS'] == month) & (origin_report['SEMANA'] == week)]
        receitas = round(temp2['ENTRADA'].sum(), 10)
        despesas = round(temp2['SAIDA'].sum(), 10)
        return round(init_balance + receitas - despesas,10)
   
    def build_weekly_report(self,fat_table: pd.DataFrame):
        report = fat_table.merge(self.TBL_Category, how='left', on='ID_Category')
        report = report.merge(self.DIM_NR_Balance, how='left')
        report = report.merge(self.DIM_NR_Income, how='left')\
                           .rename(columns={'NR_Value':'ENTRADA'})
        report = report.merge(self.DIM_NR_Expense, how='left')\
                            .rename(columns={'NR_Value':'SAIDA'})
        report = report.merge(self.DIM_DT_TransactionDate, how='left')
        report = report.sort_values(by=['DT_TransactionDate'])

        report[['ENTRADA','SAIDA']] = report[['ENTRADA','SAIDA']].fillna(0)
        translated_months = report['DT_TransactionDate'].apply(lambda d: MESES[d.month])
        report['MÊS'] = translated_months
        report['MÊS'] = pd.Categorical(report['MÊS'], translated_months.unique())
        report['ANO'] = report['DT_TransactionDate'].dt.year
        report['SEMANA'] = self._calc_week(report['DT_TransactionDate'])
        report['CATEGORIA'] = report['DS_Name'].fillna('desconhecido')
        return report
    
    def _calc_week(self, series: pd.Series) -> pd.Series:
        first_week = (series - pd.to_timedelta(series.dt.day - 1, unit='d')).dt.weekday
        return (series.dt.day - 1 + first_week) // 7 + 1

    def _load_categories(self,) -> dict:
        categorias = {}
        for row in self.TBL_Category.merge(self.TBL_Category, left_on='ID_CategoryParent',right_on='ID_Category')[['DS_Name_x','NR_Level_x','DS_Name_y']].sort_values('NR_Level_x').to_dict('records'):
            parent  = row['DS_Name_y']
            child   = row['DS_Name_x']
            if parent not in categorias:
                categorias[parent] = {}
            categorias[parent][child] = []
        return categorias

    def _load_historic(self,) -> pd.DataFrame:
        return self.TBL_DescriptionCategory.merge(self.TBL_Category)\
                        .rename(columns={'DS_Name': 'DS_Category'})[['DS_Description','DS_Category']]
    
    def import_statement(self, command:ImportationCommand):
        self.TBL_Category = self.uof.CategoryRepository.to_dataframe()
        self.TBL_DescriptionCategory = self.uof.DescriptionCategoryRepository.to_dataframe()
        self.TBL_Transactions = self.uof.TransactionsRepository.to_dataframe()
        self.errors = pd.DataFrame()
        self.duplicates = pd.DataFrame()
        self.statement_duplicates = pd.DataFrame()

        if command['bank_name'] == 'inter':
            self.tmp = adapters.bank_statement.inter.load_statement(command['statement_path'],\
                                                                    skip_rows=command['skip_rows'])
        else:
            raise Exception('Bank not implemented')
        
        ## TODO:
        ## Check if the statement is already imported: composite key: (ID_Account, DS_Description,DT_TransactionDate,NR_Balance,NR_Value,CD_Type)
        composite_key = ['ID_Account','DS_Description','DT_TransactionDate','NR_Balance','NR_Amount','CD_Type']
        self.tmp['CD_Type'] = self.tmp['NR_Value'].apply(lambda x: 'Income' if x > 0 else 'Expense')
        self.tmp['NR_Amount'] = self.tmp['NR_Value'].apply(lambda x: abs(x))
        self.tmp['ID_Account'] = pd.Series(command['id_account'], index=self.tmp.index)        
        search_duplicates = pd.concat([self.tmp[composite_key],self.TBL_Transactions[composite_key]]).reset_index()
        self.statement_duplicates = self.tmp[self.tmp[composite_key].duplicated()]
        self.duplicates = search_duplicates[search_duplicates.duplicated(subset=composite_key,keep=False)]\
                                                                 .drop_duplicates(subset=composite_key)[composite_key]

        self.duplicates = self.duplicates.merge(self.statement_duplicates, how='outer', on=composite_key, indicator='Exist')
        self.duplicates = self.duplicates[self.duplicates['Exist'] == 'left_only'].drop(columns=['Exist'])

        d = pd.merge(self.tmp,self.duplicates,on=composite_key,how='outer', indicator='Exist')
        self.tmp = d[d['Exist'] == 'left_only'].drop(columns=['Exist','NR_Value_y'])
        self.tmp.rename(columns={'NR_Value_x':'NR_Value'}, inplace=True)
        
        if self.tmp.shape[0] > 0:
            last_id  = self.TBL_Transactions.ID_Transaction.max()
            if np.isnan(last_id):
                last_id = 0
            last_id = int(last_id) + 1
            self.tmp['ID_Transaction'] = pd.Series(list(range(last_id, last_id + self.tmp.shape[0])), index=self.tmp.index)
            self.tmp['IC_Imported'] = True
            self.tmp['DT_ImportedDate'] = pd.to_datetime(datetime.now())
            self.tmp['DT_ImportedDate'] = self.tmp['DT_ImportedDate'].astype('datetime64[ns]')
            self.tmp['DT_RegistrationDate'] = pd.to_datetime(datetime.now())
            self.tmp['DT_RegistrationDate'] =self.tmp['DT_RegistrationDate'].astype('datetime64[ns]')
            self.errors = self.tmp[~self.tmp.DS_Description.isin(self.TBL_DescriptionCategory.DS_Description)]
            # I get the ID_Category from the TBL_DescriptionCategory
            self.tmp = pd.merge(self.tmp, self.TBL_DescriptionCategory, how='left', on='DS_Description')
            
            self.TBL_Transactions = pd.concat([self.TBL_Transactions, 
                                                self.tmp.drop(columns=['NR_Value'])], sort=False)
            ## validações
            # self.base_df = self.base_df.drop_duplicates()
            self.uof.TransactionsRepository._update_cells(self.TBL_Transactions)
        



        if self.errors.shape[0] > 0 or self.duplicates.shape[0] > 0 or self.statement_duplicates.shape[0] > 0:
            return {'errors': self.errors.to_dict('records'), 
                    'duplicates': self.duplicates.to_dict('records'),
                    'statement_duplicates': self.statement_duplicates.to_dict('records')}
        else:
            return None