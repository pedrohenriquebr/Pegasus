from re import T
import openpyxl
from openpyxl.worksheet.table import Table
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.colors import Color

import pandas as pd
import numpy as np
import os

import adapters.bank_statement.inter
from database.dbconnection import DbConnection
from database.schema import schema

CACHE_DIR  = './cache'
BASE_PATH = os.path.join(CACHE_DIR,'base.csv')
MESES = {
    1:'JANEIRO',
    2:'FEVEREIRO',
    3:'MARÇO',
    4:'ABRIL',
    5:'MAIO',
    6:'JUNHO',
    7:'JULHO',
    8:'AGOSTO',
    9:'SETEMBRO',
    10:'OUTUBRO',
    11:'NOVEMBRO',
    12:'DEZEMBRO'
}




class Pegasus:
    def __init__(self,db: DbConnection,debug=False):
        self.debug = debug
        self.db = db
        
        
    
    def import_statement(self, statement_path: str, bank_name: str, id_account: int) -> None:
        if bank_name == 'inter':
            self.tmp = adapters.bank_statement.inter.load_statement(statement_path)
        else:
            raise Exception('Bank not implemented')
        errors = self.tmp[~self.tmp.DS_Description.isin(self.TBL_DescriptionCategory.DS_Description)]
        
        last_id  = self.TBL_Transactions.ID_Transaction.max()
        if np.isnan(last_id):
            last_id = 0
        last_id = last_id + 1
        self.tmp['ID_Transaction'] = pd.Series(list(range(last_id, last_id + len(self.tmp))))
        self.tmp['ID_Account'] = pd.Series(id_account, index=self.tmp.index)
        self.tmp['CD_Type'] = self.tmp['NR_Value'].apply(lambda x: 'Income' if x > 0 else 'Expense')
        self.tmp['NR_Amount'] = self.tmp['NR_Value'].apply(lambda x: abs(x))
        self.tmp['IC_Imported'] = True
        self.tmp['DT_ImportedDate'] = pd.to_datetime('now')
        self.tmp['DT_ImportedDate'] = self.tmp['DT_ImportedDate'].astype('datetime64[ns]')
        self.tmp['DT_RegistrationDate'] = pd.to_datetime('now')
        self.tmp['DT_RegistrationDate'] =self.tmp['DT_RegistrationDate'].astype('datetime64[ns]')
        self.TBL_Transactions = pd.concat([self.TBL_Transactions, self.tmp.drop(['NR_Value'], axis=1)], sort=False)
        ## validações
        # self.base_df = self.base_df.drop_duplicates()
        self.db.upsert('TBL_Transactions', self.TBL_Transactions)

    def load_data(self) -> None:
        self.base_df = self.db.worksheet('DW_Base')
        self.base_df['DT_TransactionDate'] = pd.to_datetime(self.base_df['DT_TransactionDate'])
        self.base_df  = self.base_df.sort_values('DT_TransactionDate')
        self.categorias  = self.load_categories()
        self.historico_df = self.load_historic()
    
    def load_categories(self,) -> None:
        categorias = {}
        for row in self.TBL_Category.merge(self.TBL_Category, left_on='ID_CategoryParent',right_on='ID_Category')[['DS_Name_x','NR_Level_x','DS_Name_y']].sort_values('NR_Level_x').to_dict('records'):
            parent  = row['DS_Name_y']
            child   = row['DS_Name_x']
            if parent not in categorias:
                categorias[parent] = {}
            categorias[parent][child] = []
        return categorias

    def load_historic(self,) -> None:
        return self.TBL_DescriptionCategory.merge(self.TBL_Category)\
                        .rename(columns={'DS_Name': 'DS_Category'})[['DS_Description','DS_Category']]
    
    def calc_balance(self, init_balance, origin_report, pivot_table: pd.DataFrame, idx: int):
        # MÊS , SEMANA 
        month, week = pivot_table.columns[idx][1], pivot_table.columns[idx][2]

        temp2  = origin_report[(origin_report['MÊS'] == month) & (origin_report['SEMANA'] == week)]
        receitas = temp2['ENTRADA'].sum()
        despesas = temp2['SAIDA'].sum()
        return init_balance + receitas - despesas
    
    def calc_week(self, series: pd.Series) -> pd.Series:
        first_week = (series - pd.to_timedelta(series.dt.day - 1, unit='d')).dt.weekday
        return (series.dt.day - 1 + first_week) // 7 + 1
        
    def calc_income(self, report, pivot_table: pd.DataFrame, idx: int):
        # MÊS , SEMANA 
        month, week = pivot_table.columns[idx][1], pivot_table.columns[idx][2]
        temp2  = report[(report['MÊS'] == month) & (report['SEMANA'] == week)]
        return temp2['ENTRADA'].sum()

    def build_weekly_report(self,fat_table: pd.DataFrame):
        report = fat_table.merge(self.dim_historico, how='left')\
                .merge(self.dim_historico_categoria, how='left')\
                .merge(self.dim_categoria, how='left')\
                .merge(self.dim_saldo, how='left')\
                .merge(self.dim_entrada, how='left')\
                    .rename(columns={'NR_Value':'ENTRADA'})\
                .merge(self.dim_saida, how='left')\
                    .rename(columns={'NR_Value':'SAIDA'})\
                .merge(self.dim_lancamento, how='left')\
                .sort_values(by=['DT_TransactionDate'])

        report[['ENTRADA','SAIDA']] = report[['ENTRADA','SAIDA']].fillna(0)
        report['MÊS'] = report['DT_TransactionDate'].apply(lambda d: MESES[d.month])
        report['MÊS'] = pd.Categorical(report['MÊS'], MESES.values())
        report['ANO'] = report['DT_TransactionDate'].dt.year
        report['SEMANA'] = self.calc_week(report['DT_TransactionDate'])
        report['CATEGORIA'] = report['DS_Category'].fillna('desconhecido')
        return report
    
    def build_pivot_table(self, fat_table: pd.DataFrame):
        report  = self.build_weekly_report(fat_table)
        init_balance = report['NR_Balance'][0] + report['SAIDA'][0] - report['ENTRADA'][0]
        report = report[['DT_TransactionDate','ANO','MÊS','SEMANA','ENTRADA','SAIDA','CATEGORIA']]\
            .groupby(['DT_TransactionDate','ANO','MÊS','SEMANA','CATEGORIA'])\
            .sum([['ENTRADA','SAIDA']]).reset_index().drop(columns=['DT_TransactionDate'])
        filtro = report
        filtro = pd.merge(filtro, self.dim_categoria, how='left',  left_on='CATEGORIA', right_on='DS_Category')
        filtro = pd.merge(filtro, self.dim_categoria, how='left',  left_on='ID_CategoryParent', right_on='ID_Category')
        filtro['DS_CategoryParent'] = filtro['NOME_CATEGORIA_y']
        filtro =  filtro[['ANO', 'MÊS','SEMANA', 'CATEGORIA', 'DS_CategoryParent','ENTRADA', 'SAIDA']]
        filtro['NR_Value'] = filtro['ENTRADA'] + filtro['SAIDA']
        filtro['DS_CategoryParent']  = filtro['DS_CategoryParent'].fillna('Desconhecido')
        pivot_table = pd.pivot_table(filtro, values=['NR_Value'],index=['DS_CategoryParent','CATEGORIA'],columns=['MÊS','SEMANA'], aggfunc=(np.sum))

        d = filtro.groupby(['DS_CategoryParent','CATEGORIA'],as_index=True).sum().index

        groups =[]
        k = []
        for row in d:
            if row[0] not in groups:
                groups.append(row[0])
                k.append((row[0],row[0]))
                k.append((row[0],'TOTAL'))
            k.append((row[0],row[1]))

        pivot_table = pivot_table.reindex(pd.MultiIndex.from_tuples(k,names=['DS_CategoryParent','CATEGORIA']))
        pivot_table = pivot_table.reset_index()

        pivot_table.loc[-1] = ['' , 'RECEITA', *[self.calc_income(report,pivot_table,idx) for idx in range(2,pivot_table.shape[1])]]  
        pivot_table.index = pivot_table.index + 1 
        balance_values = []

        for idx in range(2,pivot_table.shape[1]):
            init_balance = self.calc_balance(init_balance, report,pivot_table,idx)
            balance_values.append(init_balance)

        pivot_table.loc[-1] = ['' , 'SALDO', *balance_values]  
        pivot_table.index = pivot_table.index + 1  
        pivot_table = pivot_table.sort_index()
        total_search = pivot_table[pivot_table['CATEGORIA'] == 'TOTAL'] 
        rows =  total_search.iterrows()

        series  = []
        for row in rows:
            cat  = row[1][0]
            label  = row[1][1]
            values = []
            for idx in range(2,pivot_table.shape[1]):
                month, week = pivot_table.columns[idx][1], pivot_table.columns[idx][2]
                values = [*values, filtro[(filtro['MÊS'] == month ) & (filtro['SEMANA'] == week ) & (filtro['DS_CategoryParent'] == cat)]['NR_Value'].sum()]
            series.append([cat, label, *values])

        pivot_table[pivot_table['CATEGORIA'] == 'TOTAL']  = pd.DataFrame(series, columns=total_search.columns, index=total_search.index)
        pivot_table = pivot_table.drop(columns=['DS_CategoryParent'])
        return pivot_table 
    
    def build_historico(self, dim_historico_categoria: pd.DataFrame, dim_historico: pd.DataFrame, dim_categoria: pd.DataFrame):
        return dim_historico_categoria\
                    .merge(dim_historico,how='left')\
                    .merge(dim_categoria, how='left')\
                    .rename(columns={'DS_Category': 'CATEGORIA'})[['DS_Description','CATEGORIA']]\
                    .fillna('desconhecido')\
                    .drop_duplicates(['DS_Description','CATEGORIA'])
    
    def transform(self):
        # Dimensions

        # dim_entrada
        self.DIM_NR_Income = self.base_df['NR_Value'].loc[self.base_df['NR_Value'] > 0]
        self.DIM_NR_Income.drop_duplicates(inplace=True)
        self.DIM_NR_Income.reset_index(drop=True,inplace=True)
        self.DIM_NR_Income = self.DIM_NR_Income.to_frame()
        self.DIM_NR_Income['ID_ENTRADA'] = self.DIM_NR_Income.index + 1
        self.DIM_NR_Income['ID_ENTRADA'] = self.DIM_NR_Income['ID_ENTRADA'].astype(np.int64)

        # dim_saida
        self.dim_saida = self.base_df['NR_Value'].loc[self.base_df['NR_Value'] < 0]
        self.dim_saida.drop_duplicates(inplace=True)
        self.dim_saida.reset_index(drop=True,inplace=True)
        self.dim_saida = self.dim_saida.to_frame()
        self.dim_saida['ID_SAIDA'] = self.dim_saida.index + 1
        self.dim_saida['ID_SAIDA'] = self.dim_saida['ID_SAIDA'].astype(np.int64)

        # dim_historico
        self.dim_historico = self.base_df['DS_Description']
        self.dim_historico.drop_duplicates(inplace=True)
        self.dim_historico.reset_index(drop=True,inplace=True)
        self.dim_historico = self.dim_historico.to_frame()
        self.dim_historico['ID_HISTORICO'] = self.dim_historico.index + 1
        self.dim_historico['ID_HISTORICO'] = self.dim_historico['ID_HISTORICO'].astype(np.int64)

        # dim_lancamento
        self.dim_lancamento = self.base_df['DT_TransactionDate']
        self.dim_lancamento.drop_duplicates(inplace=True)
        self.dim_lancamento.reset_index(drop=True,inplace=True)
        self.dim_lancamento = self.dim_lancamento.to_frame()
        self.dim_lancamento['ID_LANCAMENTO'] = self.dim_lancamento.index + 1
        self.dim_lancamento['ID_LANCAMENTO'] = self.dim_lancamento['ID_LANCAMENTO'].astype(np.int64)

        # dim_saldo
        self.dim_saldo = self.base_df['NR_Balance']
        self.dim_saldo.drop_duplicates(inplace=True)
        self.dim_saldo.reset_index(drop=True,inplace=True)
        self.dim_saldo = self.dim_saldo.to_frame()
        self.dim_saldo['ID_SALDO'] = self.dim_saldo.index + 1
        self.dim_saldo['ID_SALDO'] = self.dim_saldo['ID_SALDO'].astype(np.int64)


        self.merged = pd.merge(self.base_df,self.dim_lancamento, how='left',on='DT_TransactionDate')
        self.merged = pd.merge(self.merged,self.DIM_NR_Income, how='left',on='NR_Value')
        self.merged = pd.merge(self.merged,self.dim_saida, how='left',on='NR_Value')
        self.merged = pd.merge(self.merged,self.dim_historico, how='left',on='DS_Description')
        self.merged = pd.merge(self.merged,self.dim_saldo, how='left',on='NR_Balance')

        # # Remove columns
        self.merged.drop(columns=['DS_Description', 'NR_Value', 'DT_TransactionDate','NR_Balance'], inplace=True,errors='ignore')
        self.dim_saida['NR_Value'] = self.dim_saida['NR_Value'].abs()


        rows = find_keys(self.categorias)

        self.dim_categoria = pd.DataFrame({
            'ID_Category': pd.Series(np.arange(start=1,stop=len(rows)+1), dtype=np.int64),
            'DS_Category': pd.Series([row['cat'] for row in rows], dtype=str),
            # temp
            'DS_CategoryParent': pd.Series([row['cat_parent'] for row in rows], dtype=str),
            # temp
            'DS_Description': pd.Series([row['value']  for row in rows], dtype=str),
            'ID_CategoryParent': pd.Series([], dtype=np.int64),
            'NR_Level': pd.Series([row['level'] for row in rows], dtype=np.int64)
        })

        tmp = self.dim_categoria.merge(self.dim_categoria,how='left',left_on='DS_CategoryParent', right_on='DS_Category')
        self.dim_categoria[['ID_Category','DS_Category','ID_CategoryParent','NR_Level']] = tmp[['ID_CATEGORIA_x','NOME_CATEGORIA_x','ID_CATEGORIA_y','NR_Level_x']]
        self.dim_categoria = self.dim_categoria.drop(columns=['DS_CategoryParent','DS_Description'])
        self.dim_categoria.drop_duplicates(inplace=True)
        self.dim_categoria['ID_Category'] = self.dim_categoria['ID_Category'].astype(np.int64)
        self.dim_categoria['ID_CategoryParent'] = self.dim_categoria['ID_CategoryParent'].fillna(0)
        self.dim_categoria['ID_CategoryParent'] = self.dim_categoria['ID_CategoryParent'].astype(np.int64)

        self.dim_historico_categoria = pd.DataFrame([],columns=['ID_HISTORICO','ID_Category'])
        
        self.dim_historico_categoria = self.dim_historico\
                                        .merge(self.historico_df, how='left', on='DS_Description')\
                                        .merge(self.dim_categoria, how='left', left_on='CATEGORIA',right_on='DS_Category')\
                                        .drop_duplicates(['DS_Description','CATEGORIA'])[['ID_HISTORICO','ID_Category']]
    
        self.dim_historico_categoria = self.dim_historico_categoria[['ID_HISTORICO','ID_Category']]
        self.dim_historico_categoria['ID_Category'] = self.dim_historico_categoria['ID_Category'].fillna(0)
        self.dim_historico_categoria['ID_Category'] = self.dim_historico_categoria['ID_Category'].astype(np.int64)
        self.dim_historico_categoria['ID_HISTORICO'] = self.dim_historico_categoria['ID_HISTORICO'].fillna(0)
        self.dim_historico_categoria['ID_HISTORICO'] = self.dim_historico_categoria['ID_HISTORICO'].astype(np.int64)

        self.cat_pais = self.dim_categoria[self.dim_categoria['NR_Level'] == 1]['DS_Category'].unique()
        self.cats = self.dim_categoria[self.dim_categoria['NR_Level'] > 1]['DS_Category'].unique()

        self.report = self.build_pivot_table(self.merged)
        self.historico = self.build_historico(self.dim_historico_categoria, self.dim_historico, self.dim_categoria)
        
        self.sheets = {
            'base':self.base_df,
            'dim_entrada': self.DIM_NR_Income,
            'dim_saida':self.dim_saida,
            'dim_historico':self.dim_historico,
            'dim_saldo':self.dim_saldo,
            'fat_extrato':self.merged,
            'dim_lancamento':self.dim_lancamento,
            'dim_historico_categoria':self.dim_historico_categoria,
            'dim_categoria':self.dim_categoria,
            'report': self.report,
            'historico': self.historico
        }
        
    def export_csv(self) -> None:
        if not os.path.exists(CACHE_DIR):    
            os.mkdir(CACHE_DIR)
        print('Export DW')

        for outfile, df in self.sheets.items():
            print(f'Exporting {outfile}.csv')
            df.to_csv(os.path.join(CACHE_DIR, f'{outfile}.csv'),sep=';',index=False, header=True, )
    
    def apply_extra_styles(self,excel_file: str) -> None:
        # load workbook template.xlsx from current directory
        wb = load_workbook(excel_file)

        ws = wb['report']

        # esconde a primeira coluna
        ws.column_dimensions['A'].hidden = True

        # esconder a primeira linha
        ws.row_dimensions[1].hidden = True

        # esconder a 4ª linha
        ws.row_dimensions[4].hidden = True

        # ajustar fonte dos nomes dos meses
        for row in ws['B2:BV2']:
            for cell in row:
                cell.font = Font(size=20)

        # ajustar fonte e cor de fundo dos numeros das semanas
        for row in ws['C3:BV3']:
            for cell in row:
                # converte string em 
                cell.fill = PatternFill("solid", fgColor=Color(rgb='00305496'))
                cell.font = Font(color="ffffffff", bold=True, size=14)

        # percorre as colunas da sheet
        for col in ws.columns:
            # auto ajusta a largura da coluna
            col_name = get_column_letter(col[0].column)
            if col_name == 'B':
                ws.column_dimensions[col_name].width = 38.57
            else:
                ws.column_dimensions[col_name].auto_size = True

        # alterar zoom da worksheet para 85%
        ws.sheet_view.zoomScale = 85

        # formatar valores para o formato R$ ##,## da célula C5:BJ45
        for row in ws['C5:BV43']:
            for cell in row:
                cell.number_format = 'R$ #,##0.00'
                cell.font = Font(size=14)

        # agrupar linhas
        last_start  = 9
        last_end = 8
        for cell in ws['B9:B47']:
            # verificar se o valor da célula é NoneType
            if cell[0].value is None:
                continue
            if cell[0].value.strip() == 'TOTAL':
                continue
            
            if cell[0].value.strip() in [*self.cat_pais, 'Desconhecido']:
                ws.row_dimensions.group(last_start,last_end, hidden=True)
                last_start  = cell[0].row + 2
                last_end = last_start - 1
            else:
                last_end +=1

        ws.row_dimensions.group(last_start,last_end, hidden=True)

        # congelar a coluna B
        ws.freeze_panes = ws['C2']

        # cor das abas das worksheets
        wb['historico'].sheet_properties.tabColor = '203764'
        wb['report'].sheet_properties.tabColor = '203764'
        ws =  wb['historico']
        for row in ws['A1:B1']:
            row[0].fill = PatternFill("solid", fgColor=Color(rgb='00305496'))
            row[0].font = Font(color="ffffffff", bold=True, size=14)
            row[1].fill = PatternFill("solid", fgColor=Color(rgb='00305496'))
            row[1].font = Font(color="ffffffff", bold=True, size=14)

        wb.save(self.excel_file)
        wb.close()
        
    def export_xlsx(self, ) -> None:
        with pd.ExcelWriter(self.excel_file, mode='a',if_sheet_exists='replace') as writer:
            for name, df in self.sheets.items():
                if name == 'report':
                    self.get_report_styles(df).to_excel(writer,sheet_name=name)
                else:
                    df.to_excel(writer,sheet_name=name,index=False)

            
            for name, df in self.sheets.items():
                if name not in  ['report','historico']:
                    writer.sheets[name].sheet_state = 'hidden'
            print(f'Exporting to {self.excel_file} sheets')
        self.apply_extra_styles(self.excel_file)
    
    def _apply_styles(self, x, col, props):
        return [ props if x.name == col else 'font-size: 11pt' for v in x.index ]
    
    def get_report_styles(self,pivot_table):
        # the  Styler object dont apply styles on columns or indexes
        
        styles = {
            'TOTAL': 'background-color: #8EA9DB; color: black; font-weight: bold',
            'NR_Balance': 'font-weight: bold;background-color: white;color:black',
            'RECEITA': 'font-weight: bold;background-color: white;color:black',
            'Desconhecido': 'background-color: #305496; color: white; font-weight: bold'
        }

        return pivot_table.style.apply(lambda value, props='': np.where(pivot_table['CATEGORIA'] == 'TOTAL', props, ''), props=styles['TOTAL'], axis=0)\
            .apply(lambda value, props='': np.where(pivot_table['CATEGORIA'] == 'NR_Balance', props, ''), props=styles['NR_Balance'], axis=0)\
            .apply(lambda value, props='': np.where(pivot_table['CATEGORIA'] == 'RECEITA', props, ''), props=styles['RECEITA'], axis=0)\
            .apply(lambda value, props='': np.where(pivot_table['CATEGORIA'] == 'Desconhecido', props, ''), props=styles['Desconhecido'], axis=0)\
            .apply(lambda value, props='': np.where(pivot_table['CATEGORIA'].isin(self.cat_pais), props, ''), props='background-color: #305496; color:white; font-weight:bold', axis=0)\
            .apply(lambda value, props='': np.where(pivot_table['CATEGORIA'].isin(self.cats), props, ''), props='background-color: white; color:black; font-weight:bold', axis=0)\
            .set_properties(**{'border-color':'Black','border-width':'thin','border-style':'solid'})\
            .apply(self._apply_styles, props='font-size: 14pt', col=('CATEGORIA', '', ''),axis=0)