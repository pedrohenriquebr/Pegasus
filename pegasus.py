import openpyxl
from openpyxl.worksheet.table import Table
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.colors import Color

import pandas as pd
import numpy as np
import os

CACHE_DIR  = './cache'
BASE_PATH = os.path.join(CACHE_DIR,'base.csv')
MESES = {
    1:'JANEIRO',
    2:'FEVEREIRO',
    3:'MARÇO',
    4:'ABRIL',
    5:'MAIO',
    6:'JUNHO',
    7:'JULHO',
    8:'AGOSTO',
    9:'SETEMBRO',
    10:'OUTUBRO',
    11:'NOVEMBRO',
    12:'DEZEMBRO'
}

def find_keys(dict_):
    key_list = []
    def recursive(obj, last_parent=None,level=1):
        for key in obj.keys():
            if type(obj[key]) == dict:
                key_list.append({'cat':key, 'level':level , 'cat_parent': last_parent, 'value': None})
                recursive(obj[key], last_parent=key,level=level+1)
            else:
                a = len(obj[key])
                if a >0:
                    for value in obj[key]:
                        key_list.append({'cat': key,'cat_parent': last_parent, 'level': level, 'value': value })

                else:
                    key_list.append({'cat': key,'cat_parent': last_parent, 'level': level, 'value': None})
        
    recursive(dict_)
    return sorted(key_list, key=lambda x: x['level'])

class Pegasus:
    def __init__(self,excel_file='./template.xlsx',debug=False):
        self.excel_file = excel_file
        self.debug = debug
        self.sheets = {
            'base':None,
            'dim_entrada': None,
            'dim_saida':None,
            'dim_historico':None,
            'dim_saldo':None,
            'fat_extrato':None,
            'dim_lancamento':None,
            'dim_historico_categoria':None,
            'dim_categoria':None,
            'report':None,
            'historico': None
        }  
    
    def append_dataframe(self, statement_path: str) -> None:
        self.tmp = pd.read_csv(open(statement_path, 'r'),delimiter=';',skiprows=7,
        converters= {
            'VALOR': self._convert_brl,
            'SALDO': self._convert_brl,
            'HISTÓRICO': lambda x: x.strip()
        })
        self.tmp['DATA LANÇAMENTO'] = pd.to_datetime(self.tmp['DATA LANÇAMENTO'], format='%d/%m/%Y')
        self.tmp = self.tmp.rename(columns={'DATA LANÇAMENTO':'DATA_LANCAMENTO',
                                            'HISTÓRICO':'HISTORICO'})
        self.base_df = pd.concat([self.base_df, self.tmp])
        self.base_df = self.base_df.drop_duplicates()
        self.base_df  = self.base_df.sort_values('DATA_LANCAMENTO')

    def load_dataframe(self) -> None:

        if not os.path.exists(CACHE_DIR):
            self.base_df = pd.DataFrame({
                'DATA_LANCAMENTO': pd.Series([], dtype='datetime64[ns]'),
                'HISTORICO': pd.Series([], dtype=str),
                'VALOR': pd.Series([], dtype=np.float64),
                'SALDO': pd.Series([], dtype=np.float64),
            })
        else:
            self.base_df = pd.read_csv(BASE_PATH,sep=';')
            self.base_df['DATA_LANCAMENTO'] = pd.to_datetime(self.base_df['DATA_LANCAMENTO'])
            self.base_df  = self.base_df.sort_values('DATA_LANCAMENTO')

    def load_spreadsheet(self) -> None:
        wb = load_workbook(self.excel_file)
        sh  = [x.title for x in wb.worksheets]

        if 'categorias' not in sh:
            print('Criar worksheet "categorias"')
            exit(0)
        if 'historico' not in sh:
            print('Criar worksheet "historico"')
            exit(0)

        self.categorias  = self.load_categories(wb)
        self.historico_df = self.load_historic(wb)
    
    def load_categories(self, workbook: openpyxl.Workbook) -> None:
        # if json:
        #     self.categorias = json.load(open('./categorias.json','r', encoding='utf-8'))
        # else:
        ws = workbook['categorias']

        categorias = {}
        last_parent = None

        for row in ws.rows:
            if row[0].value is None:
                continue
            value = row[0].value.strip()
            if row[0].fill.fgColor.tint !=0.0:
                last_parent = value
                categorias[last_parent] = {}
            else:
                categorias[last_parent][value] = []
        if self.debug:
            print(f'[DEBUG]: loading categories... \n{categorias}')
            
        return categorias

    def load_historic(self, workbook: openpyxl.Workbook) -> None:
        data = list(workbook['historico'].values)
        cols = list(data[0])
        data = list(data[1:])
        historico_df = pd.DataFrame(data, columns=cols)
        historico_df['CATEGORIA'] = historico_df['CATEGORIA'].astype(str)
        historico_df['CATEGORIA'] = historico_df['CATEGORIA'].apply(lambda x: x.strip())
        historico_df['HISTORICO'] = historico_df['HISTORICO'].astype(str)
        historico_df['HISTORICO'] = historico_df['HISTORICO'].apply(lambda x: x.strip())
        return historico_df
    
    def calc_balance(self, init_balance, origin_report, pivot_table: pd.DataFrame, idx: int):
        # MÊS , SEMANA 
        month, week = pivot_table.columns[idx][1], pivot_table.columns[idx][2]

        temp2  = origin_report[(origin_report['MÊS'] == month) & (origin_report['SEMANA'] == week)]
        receitas = temp2['ENTRADA'].sum()
        despesas = temp2['SAIDA'].sum()
        return init_balance + receitas - despesas
    
    def calc_week(self, series: pd.Series) -> pd.Series:
        first_week = (series - pd.to_timedelta(series.dt.day - 1, unit='d')).dt.weekday
        return (series.dt.day - 1 + first_week) // 7 + 1
        
    def calc_income(self, report, pivot_table: pd.DataFrame, idx: int):
        # MÊS , SEMANA 
        month, week = pivot_table.columns[idx][1], pivot_table.columns[idx][2]
        temp2  = report[(report['MÊS'] == month) & (report['SEMANA'] == week)]
        return temp2['ENTRADA'].sum()

    def build_weekly_report(self,fat_table: pd.DataFrame):
        report = fat_table.merge(self.dim_historico, how='left')\
                .merge(self.dim_historico_categoria, how='left')\
                .merge(self.dim_categoria, how='left')\
                .merge(self.dim_saldo, how='left')\
                .merge(self.dim_entrada, how='left')\
                    .rename(columns={'VALOR':'ENTRADA'})\
                .merge(self.dim_saida, how='left')\
                    .rename(columns={'VALOR':'SAIDA'})\
                .merge(self.dim_lancamento, how='left')\
                .sort_values(by=['DATA_LANCAMENTO'])

        report[['ENTRADA','SAIDA']] = report[['ENTRADA','SAIDA']].fillna(0)
        report['MÊS'] = report['DATA_LANCAMENTO'].apply(lambda d: MESES[d.month])
        report['MÊS'] = pd.Categorical(report['MÊS'], MESES.values())
        report['ANO'] = report['DATA_LANCAMENTO'].dt.year
        report['SEMANA'] = self.calc_week(report['DATA_LANCAMENTO'])
        report['CATEGORIA'] = report['NOME_CATEGORIA'].fillna('desconhecido')
        return report
    
    def build_pivot_table(self, fat_table: pd.DataFrame):
        report  = self.build_weekly_report(fat_table)
        init_balance = report['SALDO'][0] + report['SAIDA'][0] - report['ENTRADA'][0]
        report = report[['DATA_LANCAMENTO','ANO','MÊS','SEMANA','ENTRADA','SAIDA','CATEGORIA']]\
            .groupby(['DATA_LANCAMENTO','ANO','MÊS','SEMANA','CATEGORIA'])\
            .sum([['ENTRADA','SAIDA']]).reset_index().drop(columns=['DATA_LANCAMENTO'])
        filtro = report
        filtro = pd.merge(filtro, self.dim_categoria, how='left',  left_on='CATEGORIA', right_on='NOME_CATEGORIA')
        filtro = pd.merge(filtro, self.dim_categoria, how='left',  left_on='ID_CATEGORIA_PAI', right_on='ID_CATEGORIA')
        filtro['CATEGORIA_PAI'] = filtro['NOME_CATEGORIA_y']
        filtro =  filtro[['ANO', 'MÊS','SEMANA', 'CATEGORIA', 'CATEGORIA_PAI','ENTRADA', 'SAIDA']]
        filtro['VALOR'] = filtro['ENTRADA'] + filtro['SAIDA']
        filtro['CATEGORIA_PAI']  = filtro['CATEGORIA_PAI'].fillna('Desconhecido')
        pivot_table = pd.pivot_table(filtro, values=['VALOR'],index=['CATEGORIA_PAI','CATEGORIA'],columns=['MÊS','SEMANA'], aggfunc=(np.sum))

        d = filtro.groupby(['CATEGORIA_PAI','CATEGORIA'],as_index=True).sum().index

        groups =[]
        k = []
        for row in d:
            if row[0] not in groups:
                groups.append(row[0])
                k.append((row[0],row[0]))
                k.append((row[0],'TOTAL'))
            k.append((row[0],row[1]))

        pivot_table = pivot_table.reindex(pd.MultiIndex.from_tuples(k,names=['CATEGORIA_PAI','CATEGORIA']))
        pivot_table = pivot_table.reset_index()

        pivot_table.loc[-1] = ['' , 'RECEITA', *[self.calc_income(report,pivot_table,idx) for idx in range(2,pivot_table.shape[1])]]  
        pivot_table.index = pivot_table.index + 1 
        balance_values = []

        for idx in range(2,pivot_table.shape[1]):
            init_balance = self.calc_balance(init_balance, report,pivot_table,idx)
            balance_values.append(init_balance)

        pivot_table.loc[-1] = ['' , 'SALDO', *balance_values]  
        pivot_table.index = pivot_table.index + 1  
        pivot_table = pivot_table.sort_index()
        total_search = pivot_table[pivot_table['CATEGORIA'] == 'TOTAL'] 
        rows =  total_search.iterrows()

        series  = []
        for row in rows:
            cat  = row[1][0]
            label  = row[1][1]
            values = []
            for idx in range(2,pivot_table.shape[1]):
                month, week = pivot_table.columns[idx][1], pivot_table.columns[idx][2]
                values = [*values, filtro[(filtro['MÊS'] == month ) & (filtro['SEMANA'] == week ) & (filtro['CATEGORIA_PAI'] == cat)]['VALOR'].sum()]
            series.append([cat, label, *values])

        pivot_table[pivot_table['CATEGORIA'] == 'TOTAL']  = pd.DataFrame(series, columns=total_search.columns, index=total_search.index)
        pivot_table = pivot_table.drop(columns=['CATEGORIA_PAI'])
        return pivot_table 
    
    def build_historico(self, dim_historico_categoria: pd.DataFrame, dim_historico: pd.DataFrame, dim_categoria: pd.DataFrame):
        return dim_historico_categoria\
                    .merge(dim_historico,how='left')\
                    .merge(dim_categoria, how='left')\
                    .rename(columns={'NOME_CATEGORIA': 'CATEGORIA'})[['HISTORICO','CATEGORIA']]\
                    .fillna('desconhecido')\
                    .drop_duplicates(['HISTORICO','CATEGORIA'])
    
    def transform(self):
        # Dimensions

        # dim_entrada
        self.dim_entrada = self.base_df['VALOR'].loc[self.base_df['VALOR'] > 0]
        self.dim_entrada.drop_duplicates(inplace=True)
        self.dim_entrada.reset_index(drop=True,inplace=True)
        self.dim_entrada = self.dim_entrada.to_frame()
        self.dim_entrada['ID_ENTRADA'] = self.dim_entrada.index + 1
        self.dim_entrada['ID_ENTRADA'] = self.dim_entrada['ID_ENTRADA'].astype(np.int64)

        # dim_saida
        self.dim_saida = self.base_df['VALOR'].loc[self.base_df['VALOR'] < 0]
        self.dim_saida.drop_duplicates(inplace=True)
        self.dim_saida.reset_index(drop=True,inplace=True)
        self.dim_saida = self.dim_saida.to_frame()
        self.dim_saida['ID_SAIDA'] = self.dim_saida.index + 1
        self.dim_saida['ID_SAIDA'] = self.dim_saida['ID_SAIDA'].astype(np.int64)

        # dim_historico
        self.dim_historico = self.base_df['HISTORICO']
        self.dim_historico.drop_duplicates(inplace=True)
        self.dim_historico.reset_index(drop=True,inplace=True)
        self.dim_historico = self.dim_historico.to_frame()
        self.dim_historico['ID_HISTORICO'] = self.dim_historico.index + 1
        self.dim_historico['ID_HISTORICO'] = self.dim_historico['ID_HISTORICO'].astype(np.int64)

        # dim_lancamento
        self.dim_lancamento = self.base_df['DATA_LANCAMENTO']
        self.dim_lancamento.drop_duplicates(inplace=True)
        self.dim_lancamento.reset_index(drop=True,inplace=True)
        self.dim_lancamento = self.dim_lancamento.to_frame()
        self.dim_lancamento['ID_LANCAMENTO'] = self.dim_lancamento.index + 1
        self.dim_lancamento['ID_LANCAMENTO'] = self.dim_lancamento['ID_LANCAMENTO'].astype(np.int64)

        # dim_saldo
        self.dim_saldo = self.base_df['SALDO']
        self.dim_saldo.drop_duplicates(inplace=True)
        self.dim_saldo.reset_index(drop=True,inplace=True)
        self.dim_saldo = self.dim_saldo.to_frame()
        self.dim_saldo['ID_SALDO'] = self.dim_saldo.index + 1
        self.dim_saldo['ID_SALDO'] = self.dim_saldo['ID_SALDO'].astype(np.int64)


        self.merged = pd.merge(self.base_df,self.dim_lancamento, how='left',on='DATA_LANCAMENTO')
        self.merged = pd.merge(self.merged,self.dim_entrada, how='left',on='VALOR')
        self.merged = pd.merge(self.merged,self.dim_saida, how='left',on='VALOR')
        self.merged = pd.merge(self.merged,self.dim_historico, how='left',on='HISTORICO')
        self.merged = pd.merge(self.merged,self.dim_saldo, how='left',on='SALDO')

        # # Remove columns
        self.merged.drop(columns=['HISTORICO', 'VALOR', 'DATA_LANCAMENTO','SALDO'], inplace=True,errors='ignore')
        self.dim_saida['VALOR'] = self.dim_saida['VALOR'].abs()


        rows = find_keys(self.categorias)

        self.dim_categoria = pd.DataFrame({
            'ID_CATEGORIA': pd.Series(np.arange(start=1,stop=len(rows)+1), dtype=np.int64),
            'NOME_CATEGORIA': pd.Series([row['cat'] for row in rows], dtype=str),
            # temp
            'CATEGORIA_PAI': pd.Series([row['cat_parent'] for row in rows], dtype=str),
            # temp
            'HISTORICO': pd.Series([row['value']  for row in rows], dtype=str),
            'ID_CATEGORIA_PAI': pd.Series([], dtype=np.int64),
            'NIVEL_HIERARQUIA': pd.Series([row['level'] for row in rows], dtype=np.int64)
        })

        tmp = self.dim_categoria.merge(self.dim_categoria,how='left',left_on='CATEGORIA_PAI', right_on='NOME_CATEGORIA')
        self.dim_categoria[['ID_CATEGORIA','NOME_CATEGORIA','ID_CATEGORIA_PAI','NIVEL_HIERARQUIA']] = tmp[['ID_CATEGORIA_x','NOME_CATEGORIA_x','ID_CATEGORIA_y','NIVEL_HIERARQUIA_x']]
        self.dim_categoria = self.dim_categoria.drop(columns=['CATEGORIA_PAI','HISTORICO'])
        self.dim_categoria.drop_duplicates(inplace=True)
        self.dim_categoria['ID_CATEGORIA'] = self.dim_categoria['ID_CATEGORIA'].astype(np.int64)
        self.dim_categoria['ID_CATEGORIA_PAI'] = self.dim_categoria['ID_CATEGORIA_PAI'].fillna(0)
        self.dim_categoria['ID_CATEGORIA_PAI'] = self.dim_categoria['ID_CATEGORIA_PAI'].astype(np.int64)

        self.dim_historico_categoria = pd.DataFrame([],columns=['ID_HISTORICO','ID_CATEGORIA'])
        
        self.dim_historico_categoria = self.dim_historico\
                                        .merge(self.historico_df, how='left', on='HISTORICO')\
                                        .merge(self.dim_categoria, how='left', left_on='CATEGORIA',right_on='NOME_CATEGORIA')\
                                        .drop_duplicates(['HISTORICO','CATEGORIA'])[['ID_HISTORICO','ID_CATEGORIA']]
    
        self.dim_historico_categoria = self.dim_historico_categoria[['ID_HISTORICO','ID_CATEGORIA']]
        self.dim_historico_categoria['ID_CATEGORIA'] = self.dim_historico_categoria['ID_CATEGORIA'].fillna(0)
        self.dim_historico_categoria['ID_CATEGORIA'] = self.dim_historico_categoria['ID_CATEGORIA'].astype(np.int64)
        self.dim_historico_categoria['ID_HISTORICO'] = self.dim_historico_categoria['ID_HISTORICO'].fillna(0)
        self.dim_historico_categoria['ID_HISTORICO'] = self.dim_historico_categoria['ID_HISTORICO'].astype(np.int64)

        self.cat_pais = self.dim_categoria[self.dim_categoria['NIVEL_HIERARQUIA'] == 1]['NOME_CATEGORIA'].unique()
        self.cats = self.dim_categoria[self.dim_categoria['NIVEL_HIERARQUIA'] > 1]['NOME_CATEGORIA'].unique()

        self.report = self.build_pivot_table(self.merged)
        self.historico = self.build_historico(self.dim_historico_categoria, self.dim_historico, self.dim_categoria)
        
        self.sheets = {
            'base':self.base_df,
            'dim_entrada': self.dim_entrada,
            'dim_saida':self.dim_saida,
            'dim_historico':self.dim_historico,
            'dim_saldo':self.dim_saldo,
            'fat_extrato':self.merged,
            'dim_lancamento':self.dim_lancamento,
            'dim_historico_categoria':self.dim_historico_categoria,
            'dim_categoria':self.dim_categoria,
            'report': self.report,
            'historico': self.historico
        }
        
    def export_csv(self) -> None:
        if not os.path.exists(CACHE_DIR):    
            os.mkdir(CACHE_DIR)
        print('Export DW')

        for outfile, df in self.sheets.items():
            print(f'Exporting {outfile}.csv')
            df.to_csv(os.path.join(CACHE_DIR, f'{outfile}.csv'),sep=';',index=False, header=True, )
    
    def apply_extra_styles(self,excel_file: str) -> None:
        # load workbook template.xlsx from current directory
        wb = load_workbook(excel_file)

        ws = wb['report']

        # esconde a primeira coluna
        ws.column_dimensions['A'].hidden = True

        # esconder a primeira linha
        ws.row_dimensions[1].hidden = True

        # esconder a 4ª linha
        ws.row_dimensions[4].hidden = True

        # ajustar fonte dos nomes dos meses
        for row in ws['B2:BV2']:
            for cell in row:
                cell.font = Font(size=20)

        # ajustar fonte e cor de fundo dos numeros das semanas
        for row in ws['C3:BV3']:
            for cell in row:
                # converte string em 
                cell.fill = PatternFill("solid", fgColor=Color(rgb='00305496'))
                cell.font = Font(color="ffffffff", bold=True, size=14)

        # percorre as colunas da sheet
        for col in ws.columns:
            # auto ajusta a largura da coluna
            col_name = get_column_letter(col[0].column)
            if col_name == 'B':
                ws.column_dimensions[col_name].width = 38.57
            else:
                ws.column_dimensions[col_name].auto_size = True

        # alterar zoom da worksheet para 85%
        ws.sheet_view.zoomScale = 85

        # formatar valores para o formato R$ ##,## da célula C5:BJ45
        for row in ws['C5:BV43']:
            for cell in row:
                cell.number_format = 'R$ #,##0.00'
                cell.font = Font(size=14)

        # agrupar linhas
        last_start  = 9
        last_end = 8
        for cell in ws['B9:B47']:
            # verificar se o valor da célula é NoneType
            if cell[0].value is None:
                continue
            if cell[0].value.strip() == 'TOTAL':
                continue
            
            if cell[0].value.strip() in [*self.cat_pais, 'Desconhecido']:
                ws.row_dimensions.group(last_start,last_end, hidden=True)
                last_start  = cell[0].row + 2
                last_end = last_start - 1
            else:
                last_end +=1

        ws.row_dimensions.group(last_start,last_end, hidden=True)

        # congelar a coluna B
        ws.freeze_panes = ws['C2']

        # cor das abas das worksheets
        wb['historico'].sheet_properties.tabColor = '203764'
        wb['report'].sheet_properties.tabColor = '203764'
        ws =  wb['historico']
        for row in ws['A1:B1']:
            row[0].fill = PatternFill("solid", fgColor=Color(rgb='00305496'))
            row[0].font = Font(color="ffffffff", bold=True, size=14)
            row[1].fill = PatternFill("solid", fgColor=Color(rgb='00305496'))
            row[1].font = Font(color="ffffffff", bold=True, size=14)

        wb.save(self.excel_file)
        wb.close()
        
    def export_xlsx(self, ) -> None:
        with pd.ExcelWriter(self.excel_file, mode='a',if_sheet_exists='replace') as writer:
            for name, df in self.sheets.items():
                if name == 'report':
                    self.get_report_styles(df).to_excel(writer,sheet_name=name)
                else:
                    df.to_excel(writer,sheet_name=name,index=False)

            
            for name, df in self.sheets.items():
                if name not in  ['report','historico']:
                    writer.sheets[name].sheet_state = 'hidden'
            print(f'Exporting to {self.excel_file} sheets')
        self.apply_extra_styles(self.excel_file)
    
    def _apply_styles(self, x, col, props):
        return [ props if x.name == col else 'font-size: 11pt' for v in x.index ]
    
    def get_report_styles(self,pivot_table):
        # the  Styler object dont apply styles on columns or indexes
        
        styles = {
            'TOTAL': 'background-color: #8EA9DB; color: black; font-weight: bold',
            'SALDO': 'font-weight: bold;background-color: white;color:black',
            'RECEITA': 'font-weight: bold;background-color: white;color:black',
            'Desconhecido': 'background-color: #305496; color: white; font-weight: bold'
        }

        return pivot_table.style.apply(lambda value, props='': np.where(pivot_table['CATEGORIA'] == 'TOTAL', props, ''), props=styles['TOTAL'], axis=0)\
            .apply(lambda value, props='': np.where(pivot_table['CATEGORIA'] == 'SALDO', props, ''), props=styles['SALDO'], axis=0)\
            .apply(lambda value, props='': np.where(pivot_table['CATEGORIA'] == 'RECEITA', props, ''), props=styles['RECEITA'], axis=0)\
            .apply(lambda value, props='': np.where(pivot_table['CATEGORIA'] == 'Desconhecido', props, ''), props=styles['Desconhecido'], axis=0)\
            .apply(lambda value, props='': np.where(pivot_table['CATEGORIA'].isin(self.cat_pais), props, ''), props='background-color: #305496; color:white; font-weight:bold', axis=0)\
            .apply(lambda value, props='': np.where(pivot_table['CATEGORIA'].isin(self.cats), props, ''), props='background-color: white; color:black; font-weight:bold', axis=0)\
            .set_properties(**{'border-color':'Black','border-width':'thin','border-style':'solid'})\
            .apply(self._apply_styles, props='font-size: 14pt', col=('CATEGORIA', '', ''),axis=0)

    def _convert_brl(self,val: str) -> float:
        value=0
        neg = bool('-' in val)
        _val= list(val.strip().replace('- ',''))
        comma = ',' in _val and _val.index(',') 
        point = '.' in _val and _val.index('.')

        if comma:
            _val[comma] = '.'
        if point:
            _val[point] = ''
        _val = ''.join(_val)
        try:
            value = float(_val[(3 if 'R$' in _val else 0):].strip().replace(',','.'))
            value = (-value if neg else value )
        except ValueError as e:
            raise ValueError(f'Não foi possível converter \'{_val}\' para float')
        return value